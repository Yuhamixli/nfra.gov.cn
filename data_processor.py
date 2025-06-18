"""
数据处理模块 - 专业版数据清理和Excel生成
参考并改进用户的 put2excel.py 代码
"""

import os
import pandas as pd
import re
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

from utils import setup_logging, ensure_directory, clean_text


class DataProcessor:
    """专业数据处理器"""
    
    def __init__(self):
        self.logger = setup_logging()
        
        # 标准化字段映射
        self.field_mapping = {
            '序号': '序号',
            '当事人名称': '当事人名称', 
            '当事人': '当事人名称',
            '主要违法违规行为': '主要违法违规行为',
            '违法违规行为': '主要违法违规行为',
            '违法违规': '主要违法违规行为',
            '行政处罚内容': '行政处罚内容',
            '处罚内容': '行政处罚内容',
            '行政处罚': '行政处罚内容',
            '行政处罚依据': '行政处罚依据',
            '处罚依据': '行政处罚依据',
            '依据': '行政处罚依据',
            '法律依据': '行政处罚依据',
            '作出决定机关': '作出决定机关',
            '决定机关': '作出决定机关',
            '机关': '作出决定机关',
            '抓取时间': '抓取时间',
            '详情链接': '详情链接',
            '标题': '标题',
            'title': '标题',
            'category': '类别',
            'page': '页码'
        }
        
        # 标准字段顺序
        self.standard_columns = [
            '序号', '当事人名称', '主要违法违规行为', 
            '行政处罚依据', '行政处罚内容', '作出决定机关', '标题',
            '类别', '页码', '抓取时间', '详情链接'
        ]
    
    def normalize_field_names(self, data: Dict) -> Dict:
        """标准化字段名称"""
        normalized_data = {}
        
        for key, value in data.items():
            # 查找映射的标准字段名
            standard_key = self.field_mapping.get(key, key)
            normalized_data[standard_key] = value
        
        return normalized_data
    
    def clean_punishment_data(self, raw_data: Dict) -> Dict:
        """清理处罚数据"""
        try:
            # 标准化字段名
            data = self.normalize_field_names(raw_data)
            
            # 清理文本内容
            cleaned_data = {}
            for key, value in data.items():
                if isinstance(value, str):
                    # 移除多余空白
                    cleaned_value = re.sub(r'\s+', ' ', value).strip()
                    
                    # 清理特殊字符
                    cleaned_value = cleaned_value.replace('\u3000', ' ')  # 全角空格
                    cleaned_value = cleaned_value.replace('\xa0', ' ')    # 不间断空格
                    
                    # 只对特定的金额字段进行金额格式化，而不是所有包含"万元"的字段
                    # 排除"行政处罚内容"等需要保留完整信息的字段
                    amount_fields = ['处罚金额', '罚款金额', '金额']  # 明确的金额字段
                    if key in amount_fields and ('万元' in cleaned_value or '元' in cleaned_value):
                        cleaned_value = self.standardize_amount(cleaned_value)
                    
                    cleaned_data[key] = cleaned_value
                else:
                    cleaned_data[key] = value
            
            # 验证必要字段
            if not cleaned_data.get('当事人名称'):
                self.logger.warning("缺少当事人名称字段")
            
            return cleaned_data
            
        except Exception as e:
            self.logger.error(f"清理数据失败: {e}")
            return raw_data
    
    def standardize_amount(self, amount_text: str) -> str:
        """标准化金额格式"""
        try:
            # 提取数字和单位
            amount_pattern = r'(\d+(?:\.\d+)?)\s*万元'
            match = re.search(amount_pattern, amount_text)
            
            if match:
                amount = float(match.group(1))
                return f"罚款{amount}万元"
            
            # 处理其他金额格式
            amount_pattern2 = r'(\d+(?:,\d+)*(?:\.\d+)?)\s*元'
            match2 = re.search(amount_pattern2, amount_text)
            
            if match2:
                amount_str = match2.group(1).replace(',', '')
                amount = float(amount_str)
                if amount >= 10000:
                    return f"罚款{amount/10000:.1f}万元"
                else:
                    return f"罚款{amount}元"
            
            return amount_text
            
        except Exception:
            return amount_text
    
    def process_category_data(self, category_data: List[Dict], category_name: str) -> List[Dict]:
        """处理分类数据"""
        processed_data = []
        
        for i, item in enumerate(category_data, 1):
            try:
                # 清理数据
                cleaned_item = self.clean_punishment_data(item)
                
                # 添加分类信息
                cleaned_item['类别'] = category_name
                
                # 确保序号字段
                if not cleaned_item.get('序号'):
                    cleaned_item['序号'] = str(i)
                
                # 处理可能的additional_records
                if 'additional_records' in cleaned_item:
                    additional = cleaned_item.pop('additional_records')
                    if isinstance(additional, list):
                        for j, additional_item in enumerate(additional):
                            additional_cleaned = self.clean_punishment_data(additional_item)
                            additional_cleaned['类别'] = category_name
                            additional_cleaned['序号'] = f"{i}-{j+1}"
                            # 继承主记录的一些信息
                            for field in ['标题', '详情链接', '抓取时间']:
                                if field not in additional_cleaned and field in cleaned_item:
                                    additional_cleaned[field] = cleaned_item[field]
                            processed_data.append(additional_cleaned)
                
                processed_data.append(cleaned_item)
                
            except Exception as e:
                self.logger.error(f"处理数据项失败: {e}")
                continue
        
        return processed_data
    
    def create_dataframe(self, processed_data: List[Dict]) -> pd.DataFrame:
        """创建标准化的DataFrame"""
        try:
            if not processed_data:
                # 返回空DataFrame但包含标准列
                return pd.DataFrame(columns=self.standard_columns)
            
            df = pd.DataFrame(processed_data)
            
            # 确保所有标准列都存在
            for col in self.standard_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # 重新排列列的顺序
            available_columns = [col for col in self.standard_columns if col in df.columns]
            extra_columns = [col for col in df.columns if col not in self.standard_columns]
            final_columns = available_columns + extra_columns
            
            df = df[final_columns]
            
            # 排序
            if '序号' in df.columns:
                try:
                    # 尝试按数字排序
                    df['_sort_key'] = df['序号'].apply(lambda x: self.extract_sort_key(str(x)))
                    df = df.sort_values('_sort_key').drop('_sort_key', axis=1)
                except Exception:
                    # 如果失败，按字符串排序
                    df = df.sort_values('序号')
            
            # 重置索引
            df = df.reset_index(drop=True)
            
            return df
            
        except Exception as e:
            self.logger.error(f"创建DataFrame失败: {e}")
            return pd.DataFrame(processed_data)
    
    def extract_sort_key(self, seq_no: str) -> tuple:
        """提取排序键"""
        try:
            # 处理 "1", "1-1", "2-3" 等格式
            if '-' in seq_no:
                parts = seq_no.split('-')
                return (int(parts[0]), int(parts[1]))
            else:
                return (int(seq_no), 0)
        except ValueError:
            return (9999, 0)  # 无法解析的序号排在最后
    
    def generate_excel_report(self, all_data: Dict[str, List[Dict]], 
                            filename: str = None, 
                            include_summary: bool = True,
                            update_master: bool = True) -> bool:
        """生成Excel报告 - 单页合并版本，支持总表更新"""
        try:
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'excel_output/金融监管总局行政处罚信息_{timestamp}.xlsx'
            
            # 确保输出目录存在
            output_dir = os.path.dirname(filename) if os.path.dirname(filename) else '.'
            ensure_directory(output_dir)
            
            # 合并所有分类数据到单个列表
            all_records = []
            
            for category, raw_data in all_data.items():
                self.logger.info(f"处理 {category} 数据...")
                
                # 处理数据
                processed_data = self.process_category_data(raw_data, category)
                
                # 添加到总列表
                all_records.extend(processed_data)
                
                self.logger.info(f"{category}: {len(processed_data)} 条记录")
            
            if not all_records:
                self.logger.warning("没有数据需要导出")
                return False
            
            # 创建合并的DataFrame
            df = self.create_merged_dataframe(all_records)
            
            # 写入当前Excel文件
            success = self.write_excel_with_hyperlinks(df, filename, include_summary, all_data, all_records)
            
            if success and update_master:
                # 更新总表
                self.update_master_excel(df)
            
            total_records = len(all_records)
            self.logger.info(f"Excel报告生成成功: {filename}")
            self.logger.info(f"总计 {total_records} 条处罚记录，合并到单个工作表")
            
            return success
            
        except Exception as e:
            self.logger.error(f"生成Excel报告失败: {e}")
            return False
    
    def write_excel_with_hyperlinks(self, df: pd.DataFrame, filename: str, 
                                   include_summary: bool, all_data: Dict, all_records_or_stats) -> bool:
        """写入Excel文件，支持超链接。all_records_or_stats可以是记录列表或统计数据列表"""
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font
            import openpyxl
            
            # 创建工作簿
            wb = Workbook()
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 创建主数据工作表
            ws_main = wb.create_sheet('行政处罚信息')
            
            # 添加数据到工作表
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_main.append(r)
            
            # 处理超链接（详情链接列）
            link_col_idx = None
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == '详情链接':
                    link_col_idx = idx
                    break
            
            if link_col_idx:
                self.logger.info(f"处理超链接，列索引: {link_col_idx}")
                hyperlink_count = 0
                
                # 从第2行开始（第1行是表头）
                for row_idx in range(2, len(df) + 2):
                    cell = ws_main.cell(row=row_idx, column=link_col_idx)
                    url = cell.value
                    if url and isinstance(url, str) and url.startswith('http'):
                        try:
                            # 创建超链接
                            cell.hyperlink = url
                            cell.value = "查看详情"  # 显示友好的文本
                            # 设置超链接样式
                            cell.font = Font(color="0000FF", underline="single")
                            hyperlink_count += 1
                        except Exception as e:
                            self.logger.debug(f"创建超链接失败: {e}")
                
                self.logger.info(f"成功创建 {hyperlink_count} 个超链接")
            else:
                self.logger.warning("未找到详情链接列")
            
            # 创建统计工作表（如果需要）
            if include_summary:
                ws_stats = wb.create_sheet('数据统计')
                
                # 检查all_records_or_stats的类型
                if isinstance(all_records_or_stats, list) and all_records_or_stats:
                    # 如果是统计数据列表（用于总表）
                    if isinstance(all_records_or_stats[0], dict) and '分类' in all_records_or_stats[0]:
                        summary_df = pd.DataFrame(all_records_or_stats)
                    else:
                        # 如果是记录列表（用于普通测试）
                        summary_data = self.generate_summary_stats(all_data, all_records_or_stats)
                        summary_df = pd.DataFrame(summary_data)
                else:
                    # 默认情况
                    summary_data = self.generate_summary_stats(all_data, [])
                    summary_df = pd.DataFrame(summary_data)
                
                for r in dataframe_to_rows(summary_df, index=False, header=True):
                    ws_stats.append(r)
            
            # 保存文件
            wb.save(filename)
            self.logger.info(f"Excel文件保存成功: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"写入Excel文件失败: {e}")
            self.logger.info("降级到普通Excel写入")
            # 降级到普通Excel写入
            try:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='行政处罚信息', index=False)
                    if include_summary:
                        # 降级模式的统计表
                        if isinstance(all_records_or_stats, list) and all_records_or_stats:
                            if isinstance(all_records_or_stats[0], dict) and '分类' in all_records_or_stats[0]:
                                summary_df = pd.DataFrame(all_records_or_stats)
                            else:
                                summary_data = self.generate_summary_stats(all_data, all_records_or_stats)
                                summary_df = pd.DataFrame(summary_data)
                        else:
                            summary_data = self.generate_summary_stats(all_data, [])
                            summary_df = pd.DataFrame(summary_data)
                        
                        summary_df.to_excel(writer, sheet_name='数据统计', index=False)
                        
                self.logger.info(f"降级模式Excel文件保存成功: {filename}")
                return True
            except Exception as e2:
                self.logger.error(f"降级Excel写入也失败: {e2}")
                return False
    
    def update_master_excel(self, new_df: pd.DataFrame) -> bool:
        """更新总表Excel文件，确保格式与测试数据表一致"""
        try:
            master_filename = 'excel_output/金融监管总局行政处罚信息_总表.xlsx'
            
            # 确保新数据的列顺序正确（与测试表一致）
            target_column_order = [
                '序号', '标题', '当事人名称', '主要违法违规行为', 
                '行政处罚依据', '行政处罚内容', '行政处罚决定书文号', 
                '作出决定机关', '发布时间', '类别', '抓取时间', '详情链接'
            ]
            
            # 重新排列新数据的列顺序
            available_cols = [col for col in target_column_order if col in new_df.columns]
            new_df = new_df[available_cols].copy()
            
            # 加载现有总表（如果存在）
            if os.path.exists(master_filename):
                self.logger.info("加载现有总表数据...")
                try:
                    existing_df = pd.read_excel(master_filename, sheet_name='行政处罚信息')
                    self.logger.info(f"现有总表记录数: {len(existing_df)}")
                    
                    # 确保现有数据也按照目标列顺序排列
                    existing_available_cols = [col for col in target_column_order if col in existing_df.columns]
                    existing_df = existing_df[existing_available_cols].copy()
                    
                except Exception as e:
                    self.logger.warning(f"读取现有总表失败: {e}，将创建新总表")
                    existing_df = pd.DataFrame()
            else:
                self.logger.info("创建新的总表...")
                existing_df = pd.DataFrame()
            
            # 记录更新前的记录数
            before_count = len(existing_df) if len(existing_df) > 0 else 0
            
            # 合并数据并去重
            if len(existing_df) > 0:
                # 合并数据
                merged_df = pd.concat([existing_df, new_df], ignore_index=True)
            else:
                merged_df = new_df.copy()
            
            # 按业务字段组合去重，避免误删不同当事人的记录
            merged_df = self.deduplicate_records(merged_df)
            
            # 记录更新后的记录数
            after_count = len(merged_df)
            new_records_count = after_count - before_count
            
            # 重新排序和编号
            merged_df = self.sort_by_publish_time(merged_df)
            merged_df['序号'] = range(1, len(merged_df) + 1)
            
            # 确保最终列顺序正确
            final_available_cols = [col for col in target_column_order if col in merged_df.columns]
            merged_df = merged_df[final_available_cols]
            
            # 生成统计数据（包含本月更新条数）
            summary_stats = self.generate_master_summary_stats(merged_df, new_records_count)
            
            # 保存总表
            success = self.write_excel_with_hyperlinks(
                merged_df, master_filename, True, {}, summary_stats
            )
            
            if success:
                self.logger.info(f"总表更新成功: {master_filename}")
                self.logger.info(f"总表当前记录数: {len(merged_df)}")
                self.logger.info(f"本次新增记录: {new_records_count}")
                return True
            else:
                self.logger.error("总表更新失败")
                return False
                
        except Exception as e:
            self.logger.error(f"更新总表失败: {e}")
            return False
    
    def deduplicate_records(self, df: pd.DataFrame) -> pd.DataFrame:
        """按业务字段组合去重，避免误删不同当事人的记录"""
        try:
            if len(df) == 0:
                return df
            
            initial_count = len(df)
            
            # 使用更精确的业务字段组合去重
            # 主要去重字段：当事人名称 + 行政处罚决定书文号 + 作出决定机关
            primary_dedup_cols = ['当事人名称', '行政处罚决定书文号', '作出决定机关']
            
            # 检查必要的列是否存在
            available_primary_cols = [col for col in primary_dedup_cols if col in df.columns]
            
            if len(available_primary_cols) >= 2:  # 至少有2个关键字段
                self.logger.info(f"使用主要业务字段去重: {available_primary_cols}")
                
                # 创建去重键
                df['_dedup_key'] = df[available_primary_cols].fillna('').astype(str).agg('|'.join, axis=1)
                
                # 去除空的去重键（所有字段都为空的记录）
                mask_non_empty = df['_dedup_key'].str.replace('|', '').str.strip() != ''
                
                # 按去重键去重，保留第一个
                df_dedup = df[mask_non_empty].drop_duplicates(subset=['_dedup_key'], keep='first')
                
                # 删除临时列
                df_dedup = df_dedup.drop('_dedup_key', axis=1)
                
            else:
                # 降级方案：使用当事人名称 + 主要违法违规行为 + 行政处罚内容
                fallback_cols = ['当事人名称', '主要违法违规行为', '行政处罚内容']
                available_fallback_cols = [col for col in fallback_cols if col in df.columns]
                
                if available_fallback_cols:
                    self.logger.info(f"使用降级去重字段: {available_fallback_cols}")
                    
                    # 创建去重键
                    df['_dedup_key'] = df[available_fallback_cols].fillna('').astype(str).agg('|'.join, axis=1)
                    
                    # 去除空的去重键
                    mask_non_empty = df['_dedup_key'].str.replace('|', '').str.strip() != ''
                    
                    # 按去重键去重
                    df_dedup = df[mask_non_empty].drop_duplicates(subset=['_dedup_key'], keep='first')
                    
                    # 删除临时列
                    df_dedup = df_dedup.drop('_dedup_key', axis=1)
                    
                else:
                    # 最后降级方案：不去重，只是移除完全空的记录
                    self.logger.warning("缺少去重所需的关键字段，跳过去重")
                    df_dedup = df.copy()
            
            final_count = len(df_dedup)
            removed_count = initial_count - final_count
            
            if removed_count > 0:
                self.logger.info(f"业务去重完成: 删除 {removed_count} 条重复记录，保留 {final_count} 条")
            else:
                self.logger.info(f"未发现重复记录，保持 {final_count} 条记录")
            
            return df_dedup.reset_index(drop=True)
            
        except Exception as e:
            self.logger.error(f"去重处理失败: {e}")
            return df
    
    def create_merged_dataframe(self, all_records: List[Dict]) -> pd.DataFrame:
        """创建合并的DataFrame，优化列顺序和字段"""
        try:
            if not all_records:
                return pd.DataFrame()
            
            # 定义标准字段顺序 - 标题放在序号后面，删除"作出决定日期"
            standard_columns = [
                '序号',
                '标题', 
                '当事人名称',
                '主要违法违规行为',
                '行政处罚依据',
                '行政处罚内容',
                '行政处罚决定书文号',
                '作出决定机关',
                '发布时间',  # 添加发布时间字段
                '类别',
                '抓取时间',
                '详情链接'  # 只保留一个链接字段
            ]
            
            # 清理和标准化数据
            cleaned_records = []
            for record in all_records:
                cleaned_record = {}
                
                # 去重处理：优先使用'详情链接'，如果没有则使用'detail_url'
                detail_url = record.get('详情链接') or record.get('detail_url', '')
                
                # 按标准顺序处理字段
                for col in standard_columns:
                    if col == '详情链接':
                        cleaned_record[col] = detail_url
                    elif col == '标题':
                        # 优先使用'标题'，如果没有则使用'title'
                        cleaned_record[col] = record.get('标题') or record.get('title', '')
                    else:
                        cleaned_record[col] = record.get(col, '')
                
                # 特殊处理：从标题提取行政处罚决定书文号（如果该字段为空）
                if not cleaned_record.get('行政处罚决定书文号') and cleaned_record.get('标题'):
                    extracted_number = self.extract_decision_number_from_title(cleaned_record['标题'])
                    if extracted_number:
                        cleaned_record['行政处罚决定书文号'] = extracted_number
                
                # 暂时保留原始序号用于调试
                cleaned_record['原始序号'] = record.get('序号', '')
                
                cleaned_records.append(cleaned_record)
            
            # 创建DataFrame
            df = pd.DataFrame(cleaned_records, columns=standard_columns + ['原始序号'])
            
            # 数据后处理
            df = df.fillna('')  # 填充空值
            
            # 按发布时间降序排列
            df = self.sort_by_publish_time(df)
            
            # 重新生成连续序号
            df['序号'] = range(1, len(df) + 1)
            
            # 移除调试用的原始序号列
            if '原始序号' in df.columns:
                df = df.drop('原始序号', axis=1)
            
            # 确保列顺序正确
            df = df[standard_columns]
            
            self.logger.info(f"合并DataFrame创建成功，共 {len(df)} 行 {len(df.columns)} 列，按发布时间降序排列")
            return df
            
        except Exception as e:
            self.logger.error(f"创建合并DataFrame失败: {e}")
            return pd.DataFrame()
    
    def sort_by_publish_time(self, df: pd.DataFrame) -> pd.DataFrame:
        """按发布时间降序排列"""
        try:
            # 处理多种时间字段
            time_columns = ['发布时间', '抓取时间']
            
            # 查找可用的时间字段
            available_time_col = None
            for col in time_columns:
                if col in df.columns and not df[col].isna().all():
                    available_time_col = col
                    break
            
            if available_time_col:
                self.logger.info(f"使用 {available_time_col} 字段进行排序")
                
                # 转换时间格式并排序
                def parse_time(time_str):
                    if not time_str or pd.isna(time_str):
                        return pd.Timestamp.min  # 空值排到最后
                    
                    time_str = str(time_str).strip()
                    
                    # 尝试多种时间格式
                    formats = [
                        '%Y-%m-%d %H:%M:%S',  # 2024-01-15 10:30:00
                        '%Y-%m-%d',           # 2024-01-15
                        '%Y/%m/%d',           # 2024/01/15
                        '%Y年%m月%d日',        # 2024年01月15日
                    ]
                    
                    for fmt in formats:
                        try:
                            return pd.to_datetime(time_str, format=fmt)
                        except:
                            continue
                    
                    # 如果都不匹配，尝试自动解析
                    try:
                        return pd.to_datetime(time_str)
                    except:
                        return pd.Timestamp.min
                
                # 创建排序键
                df['_sort_time'] = df[available_time_col].apply(parse_time)
                
                # 按时间降序排列 (最新的在前面)
                df = df.sort_values('_sort_time', ascending=False)
                
                # 删除临时排序列
                df = df.drop('_sort_time', axis=1)
                
                self.logger.info(f"已按 {available_time_col} 降序排列")
            else:
                # 如果没有时间字段，按类别和原始序号排序
                self.logger.info("未找到时间字段，按类别排序")
                df = self.sort_by_category_and_sequence(df)
            
            return df.reset_index(drop=True)
            
        except Exception as e:
            self.logger.error(f"时间排序失败: {e}")
            # 降级到类别排序
            return self.sort_by_category_and_sequence(df)
    
    def sort_by_category_and_sequence(self, df: pd.DataFrame) -> pd.DataFrame:
        """按类别和序号排序（备用排序方法）"""
        try:
            def sort_key(row):
                category_order = {'总局机关': 1, '监管局本级': 2, '监管分局本级': 3}
                category_num = category_order.get(row['类别'], 4)
                
                # 使用原始序号进行排序
                seq_no = str(row.get('原始序号', row.get('序号', ''))).strip()
                if seq_no and seq_no.isdigit():
                    return (category_num, int(seq_no))
                elif '-' in seq_no:
                    try:
                        parts = seq_no.split('-')
                        return (category_num, int(parts[0]), int(parts[1]))
                    except:
                        return (category_num, 9999, 0)
                else:
                    return (category_num, 9999, 0)
            
            # 添加排序辅助列
            df['_sort_key'] = df.apply(sort_key, axis=1)
            df = df.sort_values('_sort_key')
            df = df.drop('_sort_key', axis=1)
            
            return df
            
        except Exception as e:
            self.logger.error(f"类别排序失败: {e}")
            return df
    
    def generate_summary_stats(self, all_data: Dict[str, List[Dict]], all_records: List[Dict]) -> List[Dict]:
        """生成汇总统计数据"""
        try:
            summary_data = []
            
            # 各分类统计
            for category, raw_data in all_data.items():
                processed_count = len([r for r in all_records if r.get('类别') == category])
                summary_data.append({
                    '分类': category,
                    '记录数': processed_count,
                    '最后更新': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
            
            # 总计行
            summary_data.append({
                '分类': '总计',
                '记录数': len(all_records),
                '最后更新': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })
            
            return summary_data
            
        except Exception as e:
            self.logger.error(f"生成汇总统计失败: {e}")
            return []
    
    def validate_data_quality(self, all_data: Dict[str, List[Dict]]) -> Dict:
        """验证数据质量"""
        quality_report = {
            'total_records': 0,
            'categories': {},
            'issues': []
        }
        
        try:
            for category, data in all_data.items():
                category_stats = {
                    'total': len(data),
                    'missing_name': 0,
                    'missing_punishment': 0,
                    'missing_authority': 0,
                    'valid_records': 0
                }
                
                for item in data:
                    # 检查必要字段
                    if not item.get('当事人名称') and not item.get('当事人'):
                        category_stats['missing_name'] += 1
                    
                    if not item.get('行政处罚内容') and not item.get('处罚内容'):
                        category_stats['missing_punishment'] += 1
                    
                    if not item.get('作出决定机关') and not item.get('决定机关'):
                        category_stats['missing_authority'] += 1
                    
                    # 计算有效记录
                    if (item.get('当事人名称') or item.get('当事人')) and \
                       (item.get('行政处罚内容') or item.get('处罚内容')):
                        category_stats['valid_records'] += 1
                
                quality_report['categories'][category] = category_stats
                quality_report['total_records'] += category_stats['total']
                
                # 记录问题
                if category_stats['missing_name'] > 0:
                    quality_report['issues'].append(f"{category}: {category_stats['missing_name']} 条记录缺少当事人名称")
                
                if category_stats['missing_punishment'] > 0:
                    quality_report['issues'].append(f"{category}: {category_stats['missing_punishment']} 条记录缺少处罚内容")
            
            return quality_report
            
        except Exception as e:
            self.logger.error(f"数据质量验证失败: {e}")
            return quality_report
    
    def export_text_files(self, all_data: Dict[str, List[Dict]], 
                         output_dir: str = 'text_output') -> bool:
        """导出为文本文件 - 兼容用户原有的处理流程"""
        try:
            ensure_directory(output_dir)
            
            for category, data in all_data.items():
                processed_data = self.process_category_data(data, category)
                
                for i, item in enumerate(processed_data, 1):
                    # 生成安全的文件名
                    safe_title = re.sub(r'[<>:"/\\|?*]', '_', item.get('标题', f'{category}_{i}'))
                    filename = os.path.join(output_dir, f"{safe_title}.txt")
                    
                    # 写入文本文件
                    with open(filename, 'w', encoding='utf-8') as f:
                        f.write(f"序号\t{item.get('序号', '')}\n")
                        f.write(f"当事人名称\t{item.get('当事人名称', '')}\n")
                        f.write(f"主要违法违规行为\t{item.get('主要违法违规行为', '')}\n")
                        f.write(f"行政处罚依据\t{item.get('行政处罚依据', '')}\n")
                        f.write(f"行政处罚内容\t{item.get('行政处罚内容', '')}\n")
                        f.write(f"作出决定机关\t{item.get('作出决定机关', '')}\n")
                        f.write(f"类别\t{item.get('类别', '')}\n")
                        f.write(f"抓取时间\t{item.get('抓取时间', '')}\n")
                        f.write(f"详情链接\t{item.get('详情链接', '')}\n")
            
            self.logger.info(f"文本文件导出完成: {output_dir}")
            return True
            
        except Exception as e:
            self.logger.error(f"导出文本文件失败: {e}")
            return False
    
    def extract_decision_number_from_title(self, title: str) -> str:
        """从标题中提取行政处罚决定书文号"""
        try:
            if not title:
                return ""
            
            import re
            
            # 模式1: 括号内的内容，如"（菏金罚决字〔2025〕12号-13号）"
            pattern1 = r'[（(]([^）)]*(?:罚决字|监罚决字)[^）)]*)[）)]'
            match1 = re.search(pattern1, title)
            if match1:
                decision_number = match1.group(1).strip()
                self.logger.debug(f"从标题括号中提取决定书文号: {decision_number}")
                return decision_number
            
            # 模式2: 空格后的内容，如"商金监罚决字〔2025〕13号、14号、15号"
            # 查找包含"罚决字"的模式
            pattern2 = r'\s+([^\s]*(?:罚决字|监罚决字)[^\s]*(?:[号、，,]\s*[^\s]*)*)'
            match2 = re.search(pattern2, title)
            if match2:
                decision_number = match2.group(1).strip()
                # 清理可能的结尾标点
                decision_number = re.sub(r'[）)]*$', '', decision_number)
                self.logger.debug(f"从标题空格后提取决定书文号: {decision_number}")
                return decision_number
            
            # 模式3: 直接包含罚决字的完整模式（不在括号内）
            pattern3 = r'([^\s（(]*(?:罚决字|监罚决字)[^\s）)]*(?:[号、，,]\s*[^\s）)]*)*)'
            matches3 = re.findall(pattern3, title)
            if matches3:
                # 选择最长的匹配（通常更完整）
                decision_number = max(matches3, key=len).strip()
                if len(decision_number) > 5:  # 确保不是太短的片段
                    self.logger.debug(f"从标题直接提取决定书文号: {decision_number}")
                    return decision_number
            
            self.logger.debug(f"无法从标题提取决定书文号: {title}")
            return ""
            
        except Exception as e:
            self.logger.warning(f"提取决定书文号失败: {e}")
            return ""
    
    def generate_master_summary_stats(self, merged_df: pd.DataFrame, new_records_count: int) -> List[Dict]:
        """生成总表的统计数据，包含本月更新条数"""
        try:
            from datetime import datetime
            import calendar
            
            current_time = datetime.now()
            current_month = current_time.strftime('%Y年%m月')
            
            summary_data = []
            
            # 分类统计
            if '类别' in merged_df.columns:
                category_stats = merged_df['类别'].value_counts()
                for category, count in category_stats.items():
                    summary_data.append({
                        '分类': category,
                        '记录数': int(count),
                        '最后更新': current_time.strftime('%Y-%m-%d %H:%M:%S')
                    })
            
            # 总计行
            summary_data.append({
                '分类': '总计',
                '记录数': len(merged_df),
                '最后更新': current_time.strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # 本月更新条数
            summary_data.append({
                '分类': f'{current_month}新增',
                '记录数': new_records_count,
                '最后更新': current_time.strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # 数据质量统计
            if len(merged_df) > 0:
                # 统计完整性
                required_fields = ['当事人名称', '行政处罚内容', '作出决定机关']
                complete_records = 0
                
                for _, row in merged_df.iterrows():
                    if all(row.get(field, '') for field in required_fields):
                        complete_records += 1
                
                summary_data.append({
                    '分类': '数据完整记录',
                    '记录数': complete_records,
                    '最后更新': f'完整性: {complete_records/len(merged_df)*100:.1f}%'
                })
            
            return summary_data
            
        except Exception as e:
            self.logger.error(f"生成总表统计失败: {e}")
            return [
                {
                    '分类': '总计',
                    '记录数': len(merged_df) if len(merged_df) > 0 else 0,
                    '最后更新': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                {
                    '分类': f'{datetime.now().strftime("%Y年%m月")}新增',
                    '记录数': new_records_count,
                    '最后更新': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            ]


# 兼容原有接口的便捷函数
def process_and_save_data(crawler_data: Dict[str, List[Dict]], 
                         excel_filename: str = None,
                         include_text_export: bool = False) -> bool:
    """处理并保存爬虫数据"""
    processor = DataProcessor()
    
    try:
        # 数据质量检查
        quality_report = processor.validate_data_quality(crawler_data)
        logging.info(f"数据质量报告: 总计 {quality_report['total_records']} 条记录")
        
        if quality_report['issues']:
            logging.warning("数据质量问题:")
            for issue in quality_report['issues']:
                logging.warning(f"  - {issue}")
        
        # 生成Excel报告
        success = processor.generate_excel_report(crawler_data, excel_filename)
        
        # 可选：导出文本文件
        if include_text_export:
            processor.export_text_files(crawler_data)
        
        return success
        
    except Exception as e:
        logging.error(f"数据处理失败: {e}")
        return False 